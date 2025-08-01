import os, time, json, re
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

# === Load keywords ===
try:
    with open('keywords.json', 'r') as file:
        keywords = json.load(file)
except Exception as e:
    print("❌ Error loading keywords.json:", e)
    keywords = {}

priority = ["Governance", "Learning", "Safety", "Climate"]

URLS = {
    "Grants": "https://ngobox.org/grant_announcement_listing.php",
    "Tenders": "https://ngobox.org/rfp_eoi_listing.php"
}

HEADERS = {
    "User-Agent": "Mozilla/5.0"
}

# Setup headless browser
chrome_options = Options()
chrome_options.add_argument("--headless")
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)

def fetch_with_selenium(url):
    try:
        driver.get(url)
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "card-block"))
        )
        return driver.page_source
    except Exception as e:
        print(f"❌ Selenium failed to load: {url} — {e}")
        return ""

def extract_description_after_apply_by(soup):
    h2_tags = soup.find_all('h2', class_='card-text')
    for h2 in h2_tags:
        strong = h2.find('strong')
        if strong and 'Apply By:' in strong.text:
            desc_parts, seen_lines = [], set()
            for sibling in h2.find_all_next():
                if sibling.name == "div" and "row_section" in sibling.get("class", []):
                    b_tag = sibling.find("b")
                    if b_tag and b_tag.get_text(strip=True).lower() == "tags":
                        break
                text_content = sibling.get_text(separator='\n', strip=True)
                lines = text_content.split('\n')
                unique_lines = [line.strip() for line in lines if line.strip() and line.strip() not in seen_lines]
                if unique_lines:
                    seen_lines.update(unique_lines)
                    desc_parts.append('\n'.join(unique_lines))
            return "\n".join(desc_parts)
    return ''

def extract_how_to_apply_from_html(description):
    if not description or not isinstance(description, str):
        return "N/A"

    custom_keywords = [
        "Selection Criteria", "Evaluation & Follow-Up", "Application Guidelines", "Eligible Applicants:",
        "Scope of Work:", "Proposal Requirements", "Evaluation Criteria", "Submission Details", "Eligible Entities",
        "How to apply", "Purpose of RFP", "Proposal Guidelines", "Eligibility Criteria", "Application must include:",
        "Eligibility", "Submission of Tender:", "Technical Bid-", "Who Can Apply", "Documents Required", "Expectation:",
        "Eligibility Criterion:", "Submission terms:", "Vendor Qualifications", "To apply",
        "To know about the eligibility criteria:", "The agency's specific responsibilities include –",
        "SELCO Foundation will be responsible for:", "Partner Eligibility Criteria", "Proposal Submission Requirements",
        "Proposal Evaluation Criteria", "Eligibility Criteria for CSOs to be part of the programme:", "Pre-Bid Queries:",
        "Response to Pre-Bid Queries:", "Submission of Bid:", "Applicant Profiles:", "What we like to see in grant applications:",
        "Research that is supported by the SVRI must:", "Successful projects are most often:", "Criteria for funding:",
        "Before you begin to write your proposal, consider that IEF prefers to fund:",
        "As you prepare your budget, these are some items that IEF will not fund:", "Organizational Profile",
        "Selection Process", "Proposal Submission Guidelines", "Terms and Conditions", "Security Deposit:",
        "Facilities and Support Offered under the call for proposal:", "Prospective Consultants should demonstrate:"
    ]

    norm_keywords = [kw.lower().rstrip(":") for kw in custom_keywords]
    matched_sections = []

    segments = re.split(r'(\.\s+|\n+)', description)
    segments = [s.strip() for s in segments if s.strip() and not s.strip().startswith('.')]

    i = 0
    while i < len(segments):
        segment = segments[i]
        segment_lower = segment.lower()
        if any(kw in segment_lower for kw in norm_keywords):
            section = ["• " + segment]
            i += 1
            while i < len(segments):
                next_segment = segments[i]
                next_segment_lower = next_segment.lower()
                if any(kw in next_segment_lower for kw in norm_keywords):
                    break
                section.append(next_segment)
                i += 1
            matched_sections.append(" ".join(section))
        else:
            i += 1

    return "\n".join(matched_sections).strip() or "N/A"

def fetch_opportunities(type_name, base_url):
    listings, seen_links = [], set()
    page = 1
    MAX_PAGES = 5

    while page <= MAX_PAGES:
        url = f"{base_url}?page={page}"
        print(f"🔍 Scraping {type_name} Page {page} → {url}")

        html = fetch_with_selenium(url)
        soup = BeautifulSoup(html, 'html.parser')
        cards = soup.find_all('div', class_='card-block')
        if not cards:
            print(f"⚠️ No more cards found on {type_name} Page {page}. Stopping.")
            break

        for card in cards:
            a = card.find('a', href=True)
            if not a:
                continue
            href = a['href'].strip()
            link = href if href.startswith('http') else f"https://ngobox.org/{href.lstrip('/')}"
            title = a.get_text(strip=True)

            if link in seen_links:
                continue

            try:
                detail_html = fetch_with_selenium(link)
                detail_soup = BeautifulSoup(detail_html, 'html.parser')
            except Exception as e:
                print(f"❌ Failed to load detail page: {link} — {e}")
                continue

            deadline = 'N/A'
            for tag in detail_soup.find_all('h2', class_='card-text'):
                strong = tag.find('strong')
                if strong and 'Apply By:' in strong.text:
                    deadline = tag.get_text(strip=True).replace('Apply By:', '').strip()
                    break

            description_html = extract_description_after_apply_by(detail_soup)
            description = BeautifulSoup(description_html, 'html.parser').get_text(separator=' ', strip=True)
            how_to_apply = extract_how_to_apply_from_html(description)

            text_blob = (title + " " + description + " " + how_to_apply).lower()
            matched_verticals = []
            for vertical in priority:
                for word in keywords.get(vertical, []):
                    if re.search(r'\b' + re.escape(word.lower()) + r'\b', text_blob):
                        matched_verticals.append(vertical)
                        break

            if matched_verticals:
                listings.append({
                    "Type": type_name,
                    "Title": title,
                    "Description": description,
                    "How_to_Apply": how_to_apply,
                    "Matched_Vertical": ", ".join(matched_verticals),
                    "Deadline": deadline,
                    "Link": link
                })

            seen_links.add(link)
        page += 1
        time.sleep(1)

    return listings

def run_scraper():
    all_data = []
    for name, url in URLS.items():
        all_data.extend(fetch_opportunities(name, url))

    print("📌 Total records scraped:", len(all_data))
    if not all_data:
        print("⚠️ No data to save.")
        return

    try:
        df = pd.DataFrame(all_data)
        df = df[df['Link'].notna() & (df['Link'].str.strip() != '')]
        df['Clickable_Link'] = df.apply(
            lambda row: '=HYPERLINK("{}","{}")'.format(row['Link'], row['Title'].replace('"', '""')),
            axis=1
        )
        df['Deadline_Date'] = pd.to_datetime(df['Deadline'], format='%d %b %Y', errors='coerce')
        today = pd.Timestamp(datetime.today().date())
        df = df[df['Deadline_Date'] >= today]
        df = df.sort_values(['Deadline_Date'], na_position='last')
        df = df[['Type', 'Title', 'Description', 'How_to_Apply', 'Matched_Vertical', 'Deadline', 'Clickable_Link']]

        excel_path = 'relevant_grants.xlsx'
        df.to_excel(excel_path, index=False, engine='openpyxl')

        wb = load_workbook(excel_path)
        ws = wb.active
        for col, width in {'A': 20, 'B': 50, 'C': 80, 'D': 50, 'E': 30, 'F': 15, 'G': 50}.items():
            ws.column_dimensions[col].width = width
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.column_letter in ['C', 'D']:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                else:
                    cell.alignment = Alignment(wrap_text=False, vertical='top')
        wb.save(excel_path)
        print(f"✅ Excel saved to {excel_path}")
    except Exception as e:
        print("❌ Error saving file:", e)

    driver.quit()

if __name__ == "__main__":
    run_scraper()
